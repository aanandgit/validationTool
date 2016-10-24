'''
Created on Sep 17, 2016

@author: Hankock
'''
import os
import sys
import glob
import pprint
import win32api
import openpyxl
import xmlMethods
import uatMethods
import Project_consts
import warnings
import time

#Ignore all warnings
warnings.filterwarnings("ignore")

#if (len(sys.argv) == 2):    
    
#    pwd = win32api.GetShortPathName(str(sys.argv[1]))
    
print('Please Wait...')    
#pwd = os.getcwd()
pwd = sys.argv[1]
print('Current Directory : ' + pwd)
pwd = win32api.GetShortPathName(pwd)
if (len(pwd)!=0):    
    
    
    xml_files = []
    xlsx_files = [] 
    xml_data = {}
    xml_file_data = []
    TestCoverage = ''
    coverage_data = {}
    
    for f in glob.glob(pwd + '\*.xml'):
        xml_files.append(f)
    
    for f in glob.glob(pwd + '\*.xlsx'):
        xlsx_files.append(f)
    #pprint.pprint(xml_files)   
    
    for i in range(0, len(xlsx_files)):
        if ((xlsx_files[i].find('UAT_')) != -1):
            Source_UAT = xlsx_files[i]
        if ((xlsx_files[i].find('TestCoverage_')) != -1):
            TestCoverage = xlsx_files[i]
            
    print('Source UAT : ' + Source_UAT)
    #print('Test Coverage : ' + TestCoverage)
        
    # File Path of the generated file after parsing TS Reports and UAT
    ReportFilePath = pwd + Project_consts.Report_File_Name
    
    ## Parse source UAT to get all Test Cases
    wb_r = openpyxl.load_workbook(Source_UAT)
    uat_data = uatMethods.TestCasesInWorkbook(wb_r)
    #pprint.pprint(uat_data)
    
    ##Parse TestCOverage sheet to get list of TCs that are not to be performed
    if len(TestCoverage) != 0:
        wbc_r = openpyxl.load_workbook(TestCoverage)
        coverage_data = uatMethods.TestCasesInWorkbook(wbc_r)
     
           
    #pprint.pprint(coverage_data)
    
    ## Create ReportSummary File    
    wb_w = openpyxl.Workbook()
    wb_w.create_sheet(title = Project_consts.Sheet_name_1, index = 0)
    wb_w.create_sheet(title = Project_consts.Sheet_name_2, index = 1)
    wb_w.create_sheet(title = Project_consts.Sheet_name_3, index = 2)
    wb_w.remove_sheet(wb_w.get_sheet_by_name('Sheet'))
    wb_w.save(ReportFilePath)
        
    
    #Parse XML files     
    for i in range(0, len(xml_files)):
        xml_file_data.append(xmlMethods.DictFromXMLfile(xml_files[i]))
        xml_data = xmlMethods.mergeDictsIntoOne(xml_data,xml_file_data[i])
        #print('---- xml_file_data ----')
        #pprint.pprint(xml_files[i])
        #pprint.pprint(xml_file_data)
        #print('-----------------------')
        
    #pprint.pprint(len(xml_data))
    #pprint.pprint(xml_data)
    
    # Generate Sheets of the Report File
    xmlMethods.CreateReportSheet(wb_r, wb_w, Project_consts.Sheet_name_2, xml_data,ReportFilePath)
#    xmlMethods.CreateSummarySheet(wb_w, Project_consts.Sheet_name_1, uat_data, xml_data,ReportFilePath)
#    xmlMethods.CreateTestsNotPerformed(wb_w,Project_consts.Sheet_name_3,uat_data, xml_data,coverage_data ,ReportFilePath)
    print('Report Generated.')
    time.sleep(1)
else:
    print('Invalid Inputs, Check Inputs and try again !!')
    print('UAT should be in .xlsx format only!!')
    print('Folder Path Name should be less than 100 characters, try placing the folder on Desktop')
