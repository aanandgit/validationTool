'''
Created on Sep 10, 2016

@author: Hankock
'''
from openpyxl.styles import Font, Alignment
import pprint
def CreateSummarySheetFromData(wb_read, wb_write, Data, file_name):
   
    Sheet_write = wb_write.get_sheet_by_name('Summary')
    
    Count_TCs = TestCasesCountInAllSheets(wb_read)   
    
    ft = Font(bold=True)
    al = Alignment(horizontal='center')
    
    Sheet_write.cell(row=2, column=2).value = 'Sheet Name'
    Sheet_write.cell(row=2, column=3).value = 'No of Testcases'
    
    
    Sheet_write.cell(row=2, column=2).font = ft
    Sheet_write.cell(row=2, column=3).font = ft
    
    Sheet_write.cell(row=2, column=2).alignment = al
    Sheet_write.cell(row=2, column=3).alignment = al
    
    Sheet_write.column_dimensions['B'].width = 35
    Sheet_write.column_dimensions['C'].width = 20
    
    r = 3
    c = 2
    
    # for i in sorted(Count_TCs.items()):
    for key, value in sorted(Count_TCs.items()):
        Sheet_write.cell(row=r, column=c).value = key
        Sheet_write.cell(row=r, column=c + 1).value = value
        Sheet_write.cell(row=r, column=c + 1).alignment = al
        r += 1
        
    
    wb_write.save(file_name)
     
def CreateReportSheetFromData(wb_read, wb_write, Data, file_name):
    
    Sheet_write = wb_write.get_sheet_by_name('Report')
    
    List_of_sheets = wb_read.get_sheet_names()
        
    c = 2  # column Iterator
    k = 3  # row Iterator
    s = 1  # Sheet Iterator
    n = 0  # no of rows written
   
       
    No_of_TCs = TestCasesCountInSheet(wb_read, List_of_sheets[s])
    
    for i, j in sorted(Data.items()):  # sorts the dict before iterating
        Sheet_write.cell(row=k, column=c).value = i
        Sheet_write.cell(row=k, column=c + 1).value = j
        
        n = n + 1
        if n == No_of_TCs:
            n = 0
            c = c + 2
            k = 3
            s = s + 1 
            No_of_TCs = TestCasesCountInSheet(wb_read, List_of_sheets[s])
        else:
            k = k + 1            
    # print(Sheet_write.cell(row = i, column = 2).value + ' ' + Sheet_write.cell(row = i, column = 3).value)

    wb_write.save(file_name)

def SheetNamesbyTCnames(Workbook):
    List_of_Sheets = Workbook.get_sheet_names()
    Num_of_sheets  = len(List_of_Sheets)
    #print(List_of_Sheets)
    #print(Num_of_sheets)
    Sheet_data = {}
    Workbook_data = {}
    
    for i in range(1, Num_of_sheets):
        Sheet_data = SheetNameofTC(Workbook, List_of_Sheets[i])
        #Workbook_data.setdefault(Sheet_data.keys()[0], Sheet_data.values()[0])
        Workbook_data.update(Sheet_data)
        #print(str(Sheet_data.keys()) + ' ' + str(Sheet_data.values()))
    
    #pprint.pprint(Workbook_data)    
    return Workbook_data
    
def SheetNameofTC(Workbook, Sheet):
    Index = GetDataIndex(Workbook, Sheet, 'TestCase Name')
    column_num = Index.values()[0]
    Sheet_data = {}
    s_name = []
    flag = True
    
    Sheet_obj = Workbook.get_sheet_by_name(Sheet)    
    for i in range(1, (Sheet_obj.max_row+1)):
        row_value = str(Sheet_obj.cell(row=i, column=column_num).value)
        if (row_value.startswith('TC') == True):
            first, middle, last = row_value.partition('_')   
            s_name.append(first)
  
      
    for i in range(0, len(s_name)):
        if (s_name[0] != s_name[i]):
            flag = False
    
    if flag != False and len(s_name) != 0:
        Sheet_data.setdefault(s_name[0],Sheet)    
        
  
    return Sheet_data
    
def TestCasesInWorkbook(Workbook):
    List_of_sheets = Workbook.get_sheet_names()
    #pprint.pprint(List_of_sheets)
    Num_of_sheets = len(List_of_sheets)
    #print(Num_of_sheets)
    
    Sheet_data = {}
    Data_TCs = {}
    
    for i in range(2, Num_of_sheets-1):#Skip the last sheet - Revision History
        #pprint.pprint(List_of_sheets[i])
        Sheet_data = TestCasesInSheet(Workbook, List_of_sheets[i])
        #print(str(List_of_sheets[i]) + ' ' + str(len(Sheet_data)))
        for j in range(0, len(Sheet_data)):
            Data_TCs.setdefault(Sheet_data.keys()[j], Sheet_data.values()[j])
        
    return Data_TCs
          
def TestCasesInSheet(Workbook, Sheet):
    Index = GetDataIndex(Workbook, Sheet, 'TestCase Name')
    column_num = Index.values()[0]
    Sheet_data = {}
    Sheet = Workbook.get_sheet_by_name(Sheet)    
    for i in range(1, (Sheet.max_row+1)):
        row_value = str(Sheet.cell(row=i, column=column_num).value)
        if (row_value.startswith('TC') == True):
            Sheet_data.setdefault(row_value, 'False') 
                                    
    return Sheet_data

def TestCasesCountInAllSheets(Workbook):
    List_of_sheets = Workbook.get_sheet_names()
    Num_of_sheets = len(List_of_sheets)
    Count_TCs = {}
    
    for i in range(1, Num_of_sheets):
        Num_of_TCs = TestCasesCountInSheet(Workbook, List_of_sheets[i])
        Count_TCs.setdefault(List_of_sheets[i], Num_of_TCs)

    return Count_TCs

def TestCasesCountInSheet(Workbook, Sheet):
             
    No_of_TCs = 0
    # Get Index of Test Case Name  in the current sheet
    # Get column no of TestCase number
    Index = GetDataIndex(Workbook, Sheet, 'TestCase Name')
    column_num = Index.values()[0]
       
    Sheet = Workbook.get_sheet_by_name(Sheet)
   
    for j in range(1, Sheet.max_row):
        row_value = str(Sheet.cell(row=j, column=column_num).value)
        if (row_value.startswith('TC') == True):
            No_of_TCs = No_of_TCs + 1
       
    return No_of_TCs

def GetDataIndex(Workbook, Sheet, data):
    Index = {}
    Sheet = Workbook.get_sheet_by_name(Sheet)
    for i in range(1, Sheet.max_row):
        for j in range(1, Sheet.max_column):
            column_value = Sheet.cell(row=i, column=j).value
            # column_value and data should have same case ..???
            # print(column_value)
            if (str(column_value) == str(data)):  # does it need to be in str..???
                
                break
        break   
    Index.setdefault(i, j)  # should be integer value only
       
    return Index

#===============================================================================
# This function sets the column width
#===============================================================================
def setColumnWidth(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 1