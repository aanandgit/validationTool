'''
Created on Sep 17, 2016

@author: Hankock
'''
import xml.etree.ElementTree as ET
import pprint
import uatMethods
from openpyxl.styles import Alignment, Color
from openpyxl.styles.fills import PatternFill

#===============================================================================
# This function creates the Summary sheet. This sheet gives a summary of the 
# UAT and the Test Stand Reports.
# wb_w          : Workbook object of the file being generated.
# Sheet_name    : Name of the sheet in the workbook being written
# uat_data      : Data collected from the UAT file, TestCase Names
# xml_data      : Data collected from the XML files.
# ReportFilePath: File path location of the generated file
#===============================================================================
def CreateSummarySheet(wb_w, Sheet_name, uat_data, xml_data,ReportFilePath):
    
    Summary_sheet = wb_w.get_sheet_by_name(Sheet_name)
    
    #===========================================================================
    # Data related to the source UAT
    #===========================================================================
    
    ur = 5
    uc = 5
    
    NoOfTCsInUAT = len(uat_data)
    NoOfTCsInTS = len(xml_data)
    NoOfTCsNotAutomated = (NoOfTCsInUAT - NoOfTCsInTS)
    
    Summary_sheet.cell(row = ur, column = uc).value = 'Total No of TCs in UAT'
    Summary_sheet.cell(row = ur, column = uc+1).value = NoOfTCsInUAT
    Summary_sheet.cell(row = ur+1, column = uc).value = 'Total No of TCs in TestStand Reports'
    Summary_sheet.cell(row = ur+1, column = uc+1).value = NoOfTCsInTS
    Summary_sheet.cell(row = ur+2, column = uc).value = 'TCs Not Automted'
    Summary_sheet.cell(row = ur+2, column = uc+1).value = NoOfTCsNotAutomated
    
    
    #===========================================================================
    # Data related to the TestStand Reports
    #===========================================================================
    
    xr = 5
    xc = 2
    
    l_TCsPassFailSkip = TCsPassFailSkip(xml_data)
    
    nNoOfTCsPassed = len(l_TCsPassFailSkip[0])
    nNoOfTCsFailed = len(l_TCsPassFailSkip[1])
    nNoOfTCsSkiped = len(l_TCsPassFailSkip[2])
    nTotalNoOfTcs  = nNoOfTCsPassed + nNoOfTCsFailed + nNoOfTCsSkiped
     
    Summary_sheet.cell(row = xr, column = xc).value = 'TCs Passed'
    Summary_sheet.cell(row = xr, column = xc+1).value = nNoOfTCsPassed
    Summary_sheet.cell(row = xr+1, column = xc).value = 'TCs Failed'
    Summary_sheet.cell(row = xr+1, column = xc+1).value = nNoOfTCsFailed
    Summary_sheet.cell(row = xr+2, column = xc).value = 'TCs Skipped'
    Summary_sheet.cell(row = xr+2, column = xc+1).value = nNoOfTCsSkiped
    Summary_sheet.cell(row = xr+3, column = xc).value = 'Total No Of TestCases'
    Summary_sheet.cell(row = xr+3, column = xc+1).value = nTotalNoOfTcs
    
    
    ## Highlight cells of the there are any Skipped or Failed TCs
    if nNoOfTCsFailed != 0:
        Summary_sheet.cell(row = xr+1, column = xc).fill = PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color('FF0000'))#red
        Summary_sheet.cell(row = xr+1, column = xc+1).fill = PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color('FF0000'))#red
        
    if nNoOfTCsSkiped != 0:
        Summary_sheet.cell(row = xr+2, column = xc).fill = PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color('FFFF00'))#yellow
        Summary_sheet.cell(row = xr+2, column = xc+1).fill = PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color('FFFF00'))#yellow
    if nNoOfTCsPassed == nTotalNoOfTcs:
        Summary_sheet.cell(row = xr+0, column = xc).fill = PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color('32CD32'))#green
        Summary_sheet.cell(row = xr+0, column = xc+1).fill = PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color('32CD32'))#green
    
    uatMethods.setColumnWidth(Summary_sheet)
    wb_w.save(ReportFilePath)


#===============================================================================
# This function creates Report Sheet. This sheet lists all the Test Cases in
# the TestStand Report Files.
# wb_r              : Workbook object of UAT  
# wb_w              : Workbook object of the generated file 
# Sheet_name        : Sheet which is being written to in the generated file
# data              : Data read from XML files
# ReportFilePath    : File path of the generated file.
#===============================================================================
def CreateReportSheet(wb_r, wb_w, Sheet_name, data, ReportFilePath):
    
    r = 3   # Row Init
    c = 2   # Column Init
    
    Report_sheet = wb_w.get_sheet_by_name(Sheet_name)
    
    #names = [Sheet Name:[TestCase Names]]
    names = uatMethods.SheetNamesbyTCnames(wb_r)
    #pprint.pprint(names)
    l_testCasesInReprots = data.keys()
    l_testCaseResultsInReports = data.values()
    #pprint.pprint(l_testCasesInReprots)
    # Iterate over the list, column wise
    for i in range(1, len(names)):
        #print('names[i].keys() : ' + str(names[i].keys()))
        # List of elements i.e the value of key of this iteration
        # Since values() returns a list, index-0 of that list is the list
        # of TestCase Names needed.
        l_testCasesInSheet = names[i].values()[0]
        i_noOfTestCases = len(l_testCasesInSheet) ##.. length of list of the value of that key
        #print(l_testCasesInSheet)
        #print('len: ' + str(i_noOfTestCases))
        
        Report_sheet.cell(row = r-1, column = c).value = str(names[i].keys()[0])
        Report_sheet.merge_cells(start_row=r-1, start_column=c, end_row=r-1, end_column=c+1)
        Report_sheet.cell(row=r-1, column=c).alignment = Alignment(horizontal = 'center')
        # Iterate over the lenght of this column, i.e the length of the list
        # of TestCases of current sheet
        for j in range(0, i_noOfTestCases):    
            #print('names[i].values()[j] : ' + str(names[i].values()[j]))
            #print('data.values() : ' + str(data.keys()))
            #print(l_testCasesInSheet[j] in l_testCasesInReprots)
            if (l_testCasesInSheet[j] in l_testCasesInReprots):
                ## Found this TC in the XML file data, write it under this column 
                #pprint.pprint('match : ' + l_testCasesInSheet[j])
                Report_sheet.cell(row = r+j, column = c).value = l_testCasesInReprots[j]
                Report_sheet.cell(row = r+j, column = c+1).value = l_testCaseResultsInReports[j]
                Report_sheet.cell(row = r+j, column = c).alignment = Alignment(horizontal = 'center')
                Report_sheet.cell(row = r+j, column = c).alignment = Alignment(horizontal = 'center')
        r = 3
        c += 3
    
    '''
    wr_column_name = True
    prev_sheet_no = 1 # should be init at runtime
    curr_sheet_no = 0
    r = 3
    c = 2
    '''
        
#     Old way of creating the column of TestCase Names with Sheet Name at the top
#     for i,j in sorted(data.items()): 
#         tc_name = str(i)
#         sh_name_key, middle, last = tc_name.partition('_')
#         try : 
#             sh_name_value = names[sh_name_key]
#         except:
#             print(' Invalid TC name found XML reports : ' + tc_name)
#           
#        
#         curr_sheet_no = int(tc_name[2:5])   
#         if (curr_sheet_no != prev_sheet_no):
#             c +=3
#             r = 3
#             wr_column_name = True
#             prev_sheet_no =  curr_sheet_no
# 
#         if wr_column_name == True:
#             Report_sheet.cell(row = r-1, column = c).value = str(sh_name_value)
#             Report_sheet.merge_cells(start_row=r-1, start_column=c, end_row=r-1, end_column=c+1)
#             Report_sheet.cell(row=r-1, column=c).alignment = Alignment(horizontal = 'center')
#             wr_column_name = False
#         
#         
#         Report_sheet.cell(row = r, column=c).alignment = Alignment(horizontal = 'center')
#         Report_sheet.cell(row = r, column = c).value = i
#         Report_sheet.cell(row = r, column = c+1).value = j
#         
#         if j == 'Skipped':
#             Report_sheet.cell(row = r, column = c).fill = PatternFill(patternType='solid',
#                                         fill_type='solid', 
#                                         fgColor=Color('FFFF00'))
#             Report_sheet.cell(row = r, column = c+1).fill = PatternFill(patternType='solid',
#                                         fill_type='solid', 
#                                         fgColor=Color('FFFF00'))
# 
#         if j == 'Failed':
#             Report_sheet.cell(row = r, column = c).fill = PatternFill(patternType='solid',
#                                         fill_type='solid', 
#                                         fgColor=Color('FF0000'))
#             Report_sheet.cell(row = r, column = c+1).fill = PatternFill(patternType='solid',
#                                         fill_type='solid', 
#                                         fgColor=Color('FF0000'))
# 
#         r +=1
#     
    uatMethods.setColumnWidth(Report_sheet) 
    
    wb_w.save(ReportFilePath)    

#=======================================================================
# This function creates separate dict of Passed, Failed and Skipped 
# test cases from the TestStand Report files
#=======================================================================

def CreateTestsNotPerformed(wb_w, Sheet_name, uat_data, xml_data, coverage_data, ReportFilePath):
    
    Sheet = wb_w.get_sheet_by_name(Sheet_name)
    
    xml_list = sorted(xml_data.keys()) # init it with keys from xml_data
    not_list = sorted(uat_data.keys()) # init it with keys from uat_data
    len_not_list = len(not_list)
    len_xml_list = len(xml_list)
    uat_list = sorted(uat_data.keys()) # init it with keys from uat_data
    len_uat_list = len(uat_list)
    
    # delete the elements that are present in xml_data
    # to create the list of testcases not performed.
    for i in range(0, len_xml_list):    
        for j in range(0, len_not_list):
            if xml_list[i] == not_list[j]:
                del not_list[j]
                len_not_list -=1
                break

    not_list = sorted(not_list)
    
    len_coverage_list = len(coverage_data)
    list_coverage_data = sorted(coverage_data.keys())
    
    Sheet.freeze_panes = 'A2'        

    Sheet.column_dimensions['B'].width  = 22                  
    Sheet.column_dimensions['D'].width  = 22
    Sheet.column_dimensions['F'].width  = 22
    
    Sheet.cell(row = 1, column = 2).value = "TCs not automated"
    Sheet.cell(row = 1, column = 4).value = "TCs in TestStand Reports"
    Sheet.cell(row = 1, column = 6).value = "TCs in the UAT"
    Sheet.cell(row = 1, column = 8).value = "TestCoverage Sheet"
    
    Sheet.cell(row = 1, column = 2).alignment = Alignment(horizontal = 'center')
    Sheet.cell(row = 1, column = 4).alignment = Alignment(horizontal = 'center')
    Sheet.cell(row = 1, column = 6).alignment = Alignment(horizontal = 'center')
    Sheet.cell(row = 1, column = 8).alignment = Alignment(horizontal = 'center')
    
    for num in range(0, len(not_list)):
        Sheet.cell(row = num+3, column = 2).alignment = Alignment(horizontal = 'center')
        Sheet.cell(row = num+3, column = 2).value = not_list[num]
    for num in range(0, len(xml_data)):
        Sheet.cell(row = num+3, column = 4).alignment = Alignment(horizontal = 'center')
        Sheet.cell(row = num+3, column = 4).value = xml_list[num]
    for num in range(0, len_uat_list):
        Sheet.cell(row = num+3, column = 6).alignment = Alignment(horizontal = 'center')
        Sheet.cell(row = num+3, column = 6).value = uat_list[num]
    if (len_coverage_list != 0):
        for num in range(0, len_coverage_list):
            Sheet.cell(row = num+3, column = 8).alignment = Alignment(horizontal = 'center')
            Sheet.cell(row = num+3, column = 8).value = list_coverage_data[num]
    else:
        Sheet.cell(row = 3, column = 8).alignment = Alignment(horizontal = 'center')
        Sheet.cell(row = 3, column = 8).value = 'Test Coverage Sheet not found'    
    '''
    # Formula
    Sheet.cell(row = 3, column = 9).alignment = Alignment(horizontal = 'center')
    Sheet.cell(row = 3, column = 9).value = '=IF(F3=H3,"==","XX")
    '''

            
    #print(len(tc_list))
    #pprint.pprint(tc_list)
    
    uatMethods.setColumnWidth(Sheet)
    wb_w.save(ReportFilePath)



#===============================================================================
# Take dicts of infividual xml files and merhe them into 1 mega dict
#===============================================================================
def mergeDictsIntoOne(master_dict, new_dict):
    key_found = False
    '''
    pprint.pprint('master_dict')
    pprint.pprint(master_dict)
    pprint.pprint('new_dict')
    pprint.pprint(new_dict)
    '''
    
    if len(master_dict) == 0:
        master_dict = new_dict
    elif len(new_dict) == 0:
        new_dict = master_dict 
    elif (len(master_dict) == 0) and (len(new_dict) == 0):
        print('Could not parse xml files')
    else:
        for new_key, new_value in new_dict.items():
            for master_key, master_value in master_dict.items():
                if new_key == master_key:
                    if master_value == 'Failed':
                        # It has been marked Failed in master dict,
                        # Update it only if the status has changed to Passed
                        if new_value == 'Passed':
                            master_value = new_value
                            key_found = True
                            break
                    elif master_value == 'Skipped':
                        # It has been marked Skipped in the master dict, 
                        # Update it only if the new status is either Passed or Failed
                        if new_value != 'Skipped':
                            master_value = new_value
                            key_found = True
                            break
                
            if (key_found == False):
                #print('key not found : ' + new_key + new_value)
                master_dict.setdefault(new_key, new_value)
            elif(key_found == True):
                master_dict[master_key] = master_value
                key_found = False # Re-init it to False
    '''                
    print('Dicts merged')
    print(len(master_dict))
    pprint.pprint(master_dict) 
    '''                   
    return master_dict                
   
                
                
    
#===============================================================================
# This function creates separate lists of Pass Fail and Skipped TCs from 
# TS reports. It returns 4 lists of dicts. 
#===============================================================================
def TCsPassFailSkip(xml_data):
    passed = {}
    failed = {}
    skipped = {}
    unknown = {}
    for i,j in sorted(xml_data.items()):
        value = str(j)
        if value == 'Passed':
            passed.setdefault(i,j)            
        elif value == 'Failed':
            failed.setdefault(i,j)
        elif value == 'Skipped':
            skipped.setdefault(i,j)
        else:
            unknown.setdefault(i,j)
    return passed,failed,skipped,unknown       


#===========================================================================
# This function parse a TestStand Report file and generates a dict of all 
# the test cases and their status.
#===========================================================================
def DictFromXMLfile(file_path):
    tree = ET.parse(file_path)
    root = tree.getroot()
    
    tc_flag = False  ## Found a TC Name
    status_flag = False  ## Found a Status Value
    cfse_flag = False
    store_tc = 'NA'
    store_result = 'NA'
    result = 'NA'
    tc = 'NA'
    del_tc = []
    result_list = []
    tc_list = []
    report = {}
    
    for i in root.iter('Prop'):
        
        if i.attrib.get('TypeName', '0') == 'NI_CriticalFailureStackEntry':
            #print('NI_CriticalFailureStackEntry')
            cfse_flag = True
        
        if i.attrib.get('Name', '0') == 'BatchSerialNumber':
            cfse_flag = False
        
        if cfse_flag == True and i.attrib.get('Name', '0') == 'StepName':
            #print('CFSE_StepName')
            tc = i.find('Value')
            tc = str(tc.text)
            if tc.startswith('TC'):
                del_tc.append(tc)
                cfse_flag = False
                #print('del_tc : ' + del_tc[0])
         
                
        if i.attrib.get('Name', '0') == 'StepName':
            tc = i.find('Value')
            tc = str(tc.text)
            if tc.startswith('TC'):
                tc_flag = True  # found a TC name, should look for Status now
                store_tc = tc
          
        if i.attrib.get('Name', '0') == 'Status':
            result = i.find('Value')
            result = str(result.text)
            if result == 'Passed' or result == 'Failed' or result == 'Skipped':
                status_flag = True
                store_result = result
    
                
        if ((tc_flag == True) and (status_flag == True)):
            tc_flag = False
            status_flag = False
            #print(store_tc)
            #print(store_result)
            tc_list.append(store_tc)
            result_list.append(store_result)
        
    
    
    # Delete the TC found in NI_CriticalFailureStackEntry
    if len(del_tc) !=0:
        for i in range(0, len(del_tc)):
            for j in range(0, len(tc_list)-1):
                if(tc_list[j] == del_tc[i]):
                    result_list[j] = 'Failed'       
    
    
    '''
    print (len(tc_list))
    print(len(result_list))
    pprint.pprint(tc_list)
    pprint.pprint(result_list)
    '''
            
    for i in range(0, len(tc_list)):
        report.setdefault(tc_list[i], result_list[i])
        
    #pprint.pprint(report)
    return report
